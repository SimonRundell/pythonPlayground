# Test script: OpenPyXL
# Install OpenPyXL from the Packages button before running.
# Expected output: workbook round-trip results in the Console.
#
# IMPORTANT — file I/O limitation:
# Python's open() cannot write to the real filesystem in this browser-based
# environment.  Instead this script uses io.BytesIO to simulate a file
# in memory, which is the correct approach for Pyodide.

import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 40)
print("  OpenPyXL — functionality test")
print("=" * 40)

# --- Build workbook ---
wb = Workbook()
ws = wb.active
ws.title = "Results"

header_font  = Font(bold=True, color='FFFFFF', size=11)
header_fill  = PatternFill(fill_type='solid', fgColor='1A3A5C')
centre       = Alignment(horizontal='center', vertical='center')
thin_border  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# Headers
headers = ['Student', 'Maths', 'English', 'Science', 'Average', 'Grade']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = centre
    cell.border    = thin_border

# Data
rows = [
    ('Alice',   85, 78, 92),
    ('Bob',     72, 81, 68),
    ('Charlie', 91, 88, 95),
    ('Diana',   68, 74, 70),
    ('Eve',     95, 89, 97),
]

alt_fill = PatternFill(fill_type='solid', fgColor='EBF3FB')

for r, (name, m, en, sc) in enumerate(rows, 2):
    avg   = round((m + en + sc) / 3, 1)
    grade = 'A' if avg >= 85 else 'B' if avg >= 70 else 'C'
    data  = [name, m, en, sc, avg, grade]
    fill  = alt_fill if r % 2 == 0 else None
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.alignment = centre
        cell.border    = thin_border
        if fill:
            cell.fill = fill

# Column widths
widths = [12, 8, 9, 9, 9, 7]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# Add a second sheet
ws2 = wb.create_sheet("Summary")
ws2['A1'] = 'Summary Statistics'
ws2['A1'].font = Font(bold=True, size=12)
ws2['A2'] = 'Total students'
ws2['B2'] = len(rows)
ws2['A3'] = 'Class average'
ws2['B3'] = round(sum((m + en + sc) / 3 for _, m, en, sc in rows) / len(rows), 1)

print(f"\nWorkbook built: {len(wb.worksheets)} sheets")

# --- Round-trip through BytesIO ---
buf = io.BytesIO()
wb.save(buf)
buf.seek(0)
print(f"Serialised size: {len(buf.getvalue())} bytes")

wb2 = load_workbook(buf)
ws_read = wb2['Results']
print(f"\nRead back from memory — '{ws_read.title}' sheet contents:")
print()
for row in ws_read.iter_rows(values_only=True):
    row_str = '  '.join(str(v or '').ljust(10) for v in row)
    print(f"  {row_str}")

ws_s = wb2['Summary']
print(f"\nSummary sheet: {ws_s['A3'].value} = {ws_s['B3'].value}")

print("\n✓ OpenPyXL test complete.")
